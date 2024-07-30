import openpyxl  # pip install openpyxl


def buscar_dados_excel():
    try:
        caminho_arquivo = r"\\servidor\OBRAS\FICHA SERVICO\NOVO_MODELO_FICHA.xlsm"
        diario_pastas = "DIARIO_PASTAS"
        colunas = ["A", "B", "E", "F", "H", "T", "X"]  # Letras maiúsculas

        wb = openpyxl.load_workbook(caminho_arquivo)
        ws = wb[diario_pastas]

        # Extrair os dados das colunas, pulando a primeira linha
        diario_pastas = []
        phones = []
        enderecos = []
        instaladores = []
        recorrente = []
        dates = []
        ids_obra = []

        for coluna in colunas:
            if coluna == "A":
                for celula in ws[coluna][1:]:  # Começar a partir da segunda linha
                    ids_obra.append(celula.value)
            if coluna == "B":
                for celula in ws[coluna][1:]:  # Começar a partir da segunda linha
                    diario_pastas.append(celula.value)
            if coluna == "E":
                for celula in ws[coluna][1:]:  # Começar a partir da segunda linha
                    dates.append(celula.value.strftime("%d-%m-%y"))
            elif coluna == "F":
                for celula in ws[coluna][1:]:  # Começar a partir da segunda linha
                    enderecos.append(celula.value)
            elif coluna == "H":
                for celula in ws[coluna][1:]:  # Começar a partir da segunda linha
                    instaladores.append(celula.value)
            elif coluna == "T":
                for celula in ws[coluna][1:]:  # Começar a partir da segunda linha
                    phones.append(celula.value)
            elif coluna == "X":
                for celula in ws[coluna][1:]:  # Começar a partir da segunda linha
                    recorrente.append(celula.value)

        # Filtrar os dados para remover entradas onde o telefone é None
        filtrados = [(dp, ph, en, ins, rec, dt, obra) for dp, ph, en, ins, rec, dt, obra in
                     zip(diario_pastas, phones, enderecos, instaladores, recorrente, dates, ids_obra)
                     if ph is not None]

        if filtrados:
            (diario_pastas_filtrados, phones_filtrados, enderecos_filtrados, instaladores_filtrados,
             recorrente_filtrados, dates_filtradas, ids_filtrados) = zip(*filtrados)
        else:
            (diario_pastas_filtrados, phones_filtrados, enderecos_filtrados, instaladores_filtrados,
             recorrente_filtrados, dates_filtradas, ids_filtrados) = [], [], [], [], [], [], []

        # Retornar as listas filtradas com os dados
        return (diario_pastas_filtrados, phones_filtrados, enderecos_filtrados, instaladores_filtrados,
                recorrente_filtrados, dates_filtradas, ids_filtrados)
    except Exception as e:
        print(f"Erro ao buscar dados do Excel: {e}")
        return [], [], [], [], [], [], []


def buscar_informacao(palavra_chave):
    try:
        arquivo_excel = r"\\servidor\OBRAS\FICHA SERVICO\NOVO_MODELO_FICHA.xlsm"
        nome_planilha = "DIARIO"
        coluna_pasta = "B"
        coluna_itens = "D"

        wb = openpyxl.load_workbook(arquivo_excel)
        planilha = wb[nome_planilha]

        itens = []
        # Localiza a palavra-chave na coluna C
        for linha in planilha.iter_rows(min_row=2):  # Pula a primeira linha (cabeçalho)
            id_obra = linha[ord(coluna_pasta) - ord('A')]  # Converte letra da coluna para índice

            # Converte o valor da célula C para string antes da comparação
            valor_id_obra = str(id_obra.value).lower()  # Converte para minúsculas
            if palavra_chave.lower() in valor_id_obra:
                item = linha[ord(coluna_itens) - ord('A')].value

                itens.append(item)

        return itens
    except Exception as e:
        print(f"Erro ao buscar informação: {e}")
        return []
